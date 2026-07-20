# -*- coding: utf-8 -*-
'''
EXPOsan: Exposition of sanitation and resource recovery systems

This module is developed by:
    Ali Ahmad <aliahmad1331@gmail.com>

This module is under the University of Illinois/NCSA Open Source License.
Please refer to https://github.com/QSD-Group/EXPOsan/blob/main/LICENSE.txt
for license details.
'''
"""
Create two workbooks from the RF–TEA/LCA output:

1. Ternary_TRACI_IRR_Data_*.xlsx
   - One sheet for each of the nine TRACI categories and IRR
   - Includes only paired experimental/predicted values
   - Delta = Experimental - Predicted

2. Ternary_20to50_sfwt_summary_*.xlsx
   - Filters solids-free biocrude yield to 20–50%, inclusive
   - Includes filtered data sheets
   - Includes mean/median signed and absolute delta statistics

Solids-free yields are read directly from original_predicted_LCA.
They are not recalculated.
"""

import os
import re
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ============================================================
# USER SETTINGS
# ============================================================

INPUT_FILE = (
    r"C:\Work\Rutgers\QSDsan\EXPOsan\exposan\biobinder_ml"
    r"\results\RF_Yield_Full_TRACI_20260716_1230.xlsx"
)

COMPARISON_SHEET = "original_comparison"
PREDICTED_LCA_SHEET = "original_predicted_LCA"

TIMESTAMP = datetime.now().strftime("%Y%m%d_%H%M")
OUTPUT_DIRECTORY = os.path.dirname(INPUT_FILE)

TERNARY_OUTPUT_FILE = os.path.join(
    OUTPUT_DIRECTORY,
    f"Ternary_TRACI_IRR_Data_{TIMESTAMP}.xlsx",
)

FILTERED_OUTPUT_FILE = os.path.join(
    OUTPUT_DIRECTORY,
    f"Ternary_20to50_sfwt_summary_{TIMESTAMP}.xlsx",
)

# Inclusive filter:
BIOCRUDE_MIN = 20.0
BIOCRUDE_MAX = 50.0

# True:
#     Delta = Experimental - Predicted
#
# False:
#     Delta = Predicted - Experimental
DELTA_EXP_MINUS_PRED = True

# ============================================================
# EXCLUDE KNOWN ARTIFACTS
# ============================================================

EXCLUDED_INDICES = [227, 247]

# ============================================================
# EXACT INPUT COLUMN DEFINITIONS
# ============================================================

# Exact column names avoid:
# - Carcinogenics matching NonCarcinogenics
# - RespiratoryEffects/ParticulateMatterFormation mismatch

CATEGORY_SETTINGS = [
    {
        "sheet": "Global Warming",
        "label": "GWP",
        "unit": "kg CO2e/kg",
        "experimental":
            "Experimental GlobalWarming "
            "(kg CO2-eq/kg biobinder)",
        "predicted":
            "Predicted GlobalWarming "
            "(kg CO2-eq/kg biobinder)",
    },
    {
        "sheet": "Acidification",
        "label": "Acidification",
        "unit": "kg SO2e/kg",
        "experimental":
            "Experimental Acidification "
            "(kg SO2-eq/kg biobinder)",
        "predicted":
            "Predicted Acidification "
            "(kg SO2-eq/kg biobinder)",
    },
    {
        "sheet": "Ecotoxicity",
        "label": "Ecotoxicity",
        "unit": "CTUe/kg",
        "experimental":
            "Experimental Ecotoxicity "
            "(CTUe/kg biobinder)",
        "predicted":
            "Predicted Ecotoxicity "
            "(CTUe/kg biobinder)",
    },
    {
        "sheet": "Eutrophication",
        "label": "Eutrophication",
        "unit": "kg Ne/kg",
        "experimental":
            "Experimental Eutrophication "
            "(kg N-eq/kg biobinder)",
        "predicted":
            "Predicted Eutrophication "
            "(kg N-eq/kg biobinder)",
    },
    {
        "sheet": "Ozone Depletion",
        "label": "Ozone Depletion",
        "unit": "kg CFC-11e/kg",
        "experimental":
            "Experimental OzoneDepletion "
            "(kg CFC-11-eq/kg biobinder)",
        "predicted":
            "Predicted OzoneDepletion "
            "(kg CFC-11-eq/kg biobinder)",
    },
    {
        "sheet": "Photochemical Oxidation",
        "label": "Photochemical Oxidation",
        "unit": "kg O3e/kg",
        "experimental":
            "Experimental PhotochemicalOxidation "
            "(kg O3-eq/kg biobinder)",
        "predicted":
            "Predicted PhotochemicalOxidation "
            "(kg O3-eq/kg biobinder)",
    },
    {
        "sheet": "Carcinogenics",
        "label": "Carcinogenics",
        "unit": "CTUh/kg",
        "experimental":
            "Experimental Carcinogenics "
            "(CTUh/kg biobinder)",
        "predicted":
            "Predicted Carcinogenics "
            "(CTUh/kg biobinder)",
    },
    {
        "sheet": "Non-Carcinogenics",
        "label": "Non-Carcinogenics",
        "unit": "CTUh/kg",
        "experimental":
            "Experimental NonCarcinogenics "
            "(CTUh/kg biobinder)",
        "predicted":
            "Predicted NonCarcinogenics "
            "(CTUh/kg biobinder)",
    },
    {
        "sheet": "PM Formation",
        "label": "PM Formation",
        "unit": "kg PM2.5e/kg",
        "experimental":
            "Experimental ParticulateMatterFormation "
            "(kg PM2.5-eq/kg biobinder)",
        "predicted":
            "Predicted ParticulateMatterFormation "
            "(kg PM2.5-eq/kg biobinder)",
    },
]


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def normalize_text(value):
    """Normalize text for flexible column matching."""
    return re.sub(
        r"[^a-z0-9]+",
        "",
        str(value).lower(),
    )


def require_columns(dataframe, required_columns, table_name):
    """Raise a readable error if required columns are absent."""
    missing = [
        column
        for column in required_columns
        if column not in dataframe.columns
    ]

    if missing:
        raise KeyError(
            f"Missing required columns in {table_name}:\n"
            + "\n".join(f"  - {column}" for column in missing)
        )


def find_existing_column(columns, alternatives):
    """
    Find one existing column from a set of acceptable names.

    Exact names are checked first. Normalized names are checked second.
    """
    columns = list(columns)

    for alternative in alternatives:
        if alternative in columns:
            return alternative

    normalized_columns = {
        normalize_text(column): column
        for column in columns
    }

    for alternative in alternatives:
        normalized = normalize_text(alternative)

        if normalized in normalized_columns:
            return normalized_columns[normalized]

    raise KeyError(
        "None of the acceptable column names were found:\n"
        + "\n".join(f"  - {name}" for name in alternatives)
    )


def prepare_solids_free_yields(predicted_lca):
    """
    Read existing solids-free yields directly from the input workbook.

    No yield recalculation is performed.
    """
    biocrude_column = find_existing_column(
        predicted_lca.columns,
        [
            "Biocrude solids-free wt%",
            "Biocrude (sfwt%)",
            "Biocrude SF wt%",
        ],
    )

    aqueous_column = find_existing_column(
        predicted_lca.columns,
        [
            "Aqueous solids-free wt%",
            "Aqueous (sfwt%)",
            "Aqueous SF wt%",
        ],
    )

    gas_column = find_existing_column(
        predicted_lca.columns,
        [
            "Gas solids-free wt%",
            "Gas (sfwt%)",
            "Gas SF wt%",
        ],
    )

    require_columns(
        predicted_lca,
        ["Config", "Index"],
        PREDICTED_LCA_SHEET,
    )

    solids_free = predicted_lca[
        [
            "Config",
            "Index",
            biocrude_column,
            aqueous_column,
            gas_column,
        ]
    ].copy()

    solids_free = solids_free.rename(
        columns={
            biocrude_column: "Biocrude (sfwt%)",
            aqueous_column: "Aqueous (sfwt%)",
            gas_column: "Gas (sfwt%)",
        }
    )

    for column in [
        "Biocrude (sfwt%)",
        "Aqueous (sfwt%)",
        "Gas (sfwt%)",
    ]:
        solids_free[column] = pd.to_numeric(
            solids_free[column],
            errors="coerce",
        )

    # Diagnostic only; this does not recalculate any yields.
    solids_free["SF Yield Sum (%)"] = (
        solids_free["Biocrude (sfwt%)"]
        + solids_free["Aqueous (sfwt%)"]
        + solids_free["Gas (sfwt%)"]
    )

    duplicate_keys = solids_free.duplicated(
        subset=["Config", "Index"],
        keep=False,
    )

    if duplicate_keys.any():
        duplicate_rows = solids_free.loc[
            duplicate_keys,
            ["Config", "Index"],
        ]

        raise ValueError(
            "Duplicate Config/Index combinations were found in "
            f"{PREDICTED_LCA_SHEET}:\n"
            f"{duplicate_rows.head(20)}"
        )

    return solids_free


def create_metric_sheet(
    base,
    experimental_column,
    predicted_column,
    metric_label,
    metric_unit,
):
    """Create one paired experimental/predicted metric table."""
    require_columns(
        base,
        [
            "Config",
            "Index",
            "Biocrude (sfwt%)",
            "Aqueous (sfwt%)",
            "Gas (sfwt%)",
            "SF Yield Sum (%)",
            experimental_column,
            predicted_column,
        ],
        "merged comparison table",
    )

    experimental_name = (
        f"Experimental {metric_label} ({metric_unit})"
    )

    predicted_name = (
        f"Predicted {metric_label} ({metric_unit})"
    )

    delta_name = (
        f"Δ{metric_label} ({metric_unit})"
    )

    absolute_delta_name = (
        f"|Δ{metric_label}| ({metric_unit})"
    )

    output = base[
        [
            "Config",
            "Index",
            "Biocrude (sfwt%)",
            "Aqueous (sfwt%)",
            "Gas (sfwt%)",
            "SF Yield Sum (%)",
            experimental_column,
            predicted_column,
        ]
    ].copy()

    output = output.rename(
        columns={
            experimental_column: experimental_name,
            predicted_column: predicted_name,
        }
    )

    output[experimental_name] = pd.to_numeric(
        output[experimental_name],
        errors="coerce",
    )

    output[predicted_name] = pd.to_numeric(
        output[predicted_name],
        errors="coerce",
    )

    if DELTA_EXP_MINUS_PRED:
        output[delta_name] = (
            output[experimental_name]
            - output[predicted_name]
        )
    else:
        output[delta_name] = (
            output[predicted_name]
            - output[experimental_name]
        )

    output[absolute_delta_name] = (
        output[delta_name].abs()
    )

    # Keep only rows with:
    # - successful experimental metric
    # - successful predicted metric
    # - valid solids-free compositions
    valid_columns = [
        experimental_name,
        predicted_name,
        "Biocrude (sfwt%)",
        "Aqueous (sfwt%)",
        "Gas (sfwt%)",
    ]

    valid_mask = np.isfinite(
        output[valid_columns].to_numpy(dtype=float)
    ).all(axis=1)

    output = output.loc[valid_mask].copy()

    output = output.sort_values(
        ["Config", "Index"]
    ).reset_index(drop=True)

    return output


def identify_metric_columns(table):
    """Identify experimental, predicted, and signed delta columns."""
    experimental_columns = [
        column
        for column in table.columns
        if str(column).startswith("Experimental ")
    ]

    predicted_columns = [
        column
        for column in table.columns
        if str(column).startswith("Predicted ")
    ]

    delta_columns = [
        column
        for column in table.columns
        if str(column).startswith("Δ")
    ]

    if len(experimental_columns) != 1:
        raise KeyError(
            "Expected one experimental metric column, found:\n"
            f"{experimental_columns}"
        )

    if len(predicted_columns) != 1:
        raise KeyError(
            "Expected one predicted metric column, found:\n"
            f"{predicted_columns}"
        )

    if len(delta_columns) != 1:
        raise KeyError(
            "Expected one signed delta column, found:\n"
            f"{delta_columns}"
        )

    return (
        experimental_columns[0],
        predicted_columns[0],
        delta_columns[0],
    )


def summarize_metric_table(metric_name, table):
    """Calculate paired-data accuracy and correlation statistics."""
    experimental_column, predicted_column, delta_column = (
        identify_metric_columns(table)
    )

    paired = table[
        [
            experimental_column,
            predicted_column,
            delta_column,
        ]
    ].copy()

    paired = paired.replace(
        [np.inf, -np.inf],
        np.nan,
    ).dropna()

    if len(paired) >= 2:
        pearson = paired[
            [experimental_column, predicted_column]
        ].corr(
            method="pearson"
        ).iloc[0, 1]

        spearman = paired[
            [experimental_column, predicted_column]
        ].corr(
            method="spearman"
        ).iloc[0, 1]
    else:
        pearson = np.nan
        spearman = np.nan

    delta = paired[delta_column]

    relative_metrics = calculate_relative_error_metrics(
        table=paired,
        experimental_column=experimental_column,
        predicted_column=predicted_column,
    )

    return {
        "Metric": metric_name,
        "Valid paired rows": len(paired),

        "Experimental median":
            paired[experimental_column].median(),

        "Predicted median":
            paired[predicted_column].median(),

        "Median experimental magnitude":
            relative_metrics[
                "Median experimental magnitude"
            ],

        "Mean Δ":
            delta.mean(),

        "Median Δ":
            delta.median(),

        "Mean |Δ|":
            delta.abs().mean(),

        "Median |Δ|":
            delta.abs().median(),

        "NMAE (%)":
            relative_metrics["NMAE (%)"],

        "MdAPE (%)":
            relative_metrics["MdAPE (%)"],

        "MdAPE valid rows":
            relative_metrics["MdAPE valid rows"],

        "RMSE":
            np.sqrt(
                np.mean(delta**2)
            ),

        "Pearson r":
            pearson,

        "Spearman rho":
            spearman,
    }
        
def summarize_filtered_table(metric_name, table):
    """
    Summarize prediction errors for the selected solids-free
    biocrude-yield range.
    """
    experimental_column, predicted_column, delta_column = (
        identify_metric_columns(table)
    )

    working = table[
        [
            experimental_column,
            predicted_column,
            delta_column,
        ]
    ].copy()

    working = working.replace(
        [np.inf, -np.inf],
        np.nan,
    ).dropna()

    delta = working[delta_column]

    relative_metrics = calculate_relative_error_metrics(
        table=working,
        experimental_column=experimental_column,
        predicted_column=predicted_column,
    )

    return {
        "Metric": metric_name,

        "Biocrude minimum (sfwt%)":
            BIOCRUDE_MIN,

        "Biocrude maximum (sfwt%)":
            BIOCRUDE_MAX,

        "Valid samples":
            len(delta),

        "Experimental median":
            working[experimental_column].median(),

        "Predicted median":
            working[predicted_column].median(),

        "Median experimental magnitude":
            relative_metrics[
                "Median experimental magnitude"
            ],

        "Mean Δ":
            delta.mean(),

        "Median Δ":
            delta.median(),

        "Mean |Δ|":
            delta.abs().mean(),

        "Median |Δ|":
            delta.abs().median(),

        "NMAE (%)":
            relative_metrics["NMAE (%)"],

        "MdAPE (%)":
            relative_metrics["MdAPE (%)"],

        "MdAPE valid rows":
            relative_metrics["MdAPE valid rows"],

        "Standard deviation Δ":
            delta.std(ddof=1),

        "Minimum Δ":
            delta.min(),

        "25th percentile Δ":
            delta.quantile(0.25),

        "75th percentile Δ":
            delta.quantile(0.75),

        "Maximum Δ":
            delta.max(),
    }

def calculate_relative_error_metrics(
    table,
    experimental_column,
    predicted_column,
    near_zero_fraction=1e-6,
):
    """
    Calculate scale-normalized error metrics.

    Metrics
    -------
    NMAE (%)
        median(|Experimental - Predicted|)
        divided by median(|Experimental|), multiplied by 100.

    MdAPE (%)
        median(
            |Experimental - Predicted|
            divided by |Experimental|
        ) multiplied by 100.

    Near-zero experimental values are excluded from MdAPE because
    percentage errors become unstable when the denominator is close
    to zero.
    """
    paired = table[
        [experimental_column, predicted_column]
    ].copy()

    paired[experimental_column] = pd.to_numeric(
        paired[experimental_column],
        errors="coerce",
    )

    paired[predicted_column] = pd.to_numeric(
        paired[predicted_column],
        errors="coerce",
    )

    paired = paired.replace(
        [np.inf, -np.inf],
        np.nan,
    ).dropna()

    if paired.empty:
        return {
            "Median experimental magnitude": np.nan,
            "Median absolute error": np.nan,
            "NMAE (%)": np.nan,
            "MdAPE (%)": np.nan,
            "MdAPE valid rows": 0,
        }

    experimental = paired[experimental_column]
    predicted = paired[predicted_column]

    absolute_error = (
        experimental - predicted
    ).abs()

    median_experimental_magnitude = (
        experimental.abs().median()
    )

    median_absolute_error = (
        absolute_error.median()
    )

    if (
        np.isfinite(median_experimental_magnitude)
        and median_experimental_magnitude > 0
    ):
        nmae = (
            median_absolute_error
            / median_experimental_magnitude
            * 100
        )
    else:
        nmae = np.nan

    # Scale the near-zero cutoff to the typical magnitude of the
    # experimental metric.
    if (
        np.isfinite(median_experimental_magnitude)
        and median_experimental_magnitude > 0
    ):
        denominator_cutoff = (
            near_zero_fraction
            * median_experimental_magnitude
        )
    else:
        denominator_cutoff = 1e-12

    valid_percentage_mask = (
        experimental.abs() > denominator_cutoff
    )

    percentage_error = (
        absolute_error.loc[valid_percentage_mask]
        / experimental.loc[valid_percentage_mask].abs()
        * 100
    )

    mdape = (
        percentage_error.median()
        if not percentage_error.empty
        else np.nan
    )

    return {
        "Median experimental magnitude":
            median_experimental_magnitude,

        "Median absolute error":
            median_absolute_error,

        "NMAE (%)":
            nmae,

        "MdAPE (%)":
            mdape,

        "MdAPE valid rows":
            int(valid_percentage_mask.sum()),
    }
        
def format_excel(output_file):
    """Apply readable formatting to every workbook sheet."""
    workbook = load_workbook(output_file)

    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        for cell in worksheet[1]:
            cell.font = Font(bold=True)

        for column_cells in worksheet.columns:
            column_letter = get_column_letter(
                column_cells[0].column
            )

            max_length = max(
                len(str(cell.value))
                if cell.value is not None
                else 0
                for cell in column_cells
            )

            worksheet.column_dimensions[
                column_letter
            ].width = min(max_length + 2, 42)

        for row in worksheet.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, float):
                    cell.number_format = "0.00000E+00"

    workbook.save(output_file)


# ============================================================
# READ INPUT WORKBOOK
# ============================================================

if not os.path.isfile(INPUT_FILE):
    raise FileNotFoundError(
        f"Input workbook not found:\n{INPUT_FILE}"
    )

print("=" * 80)
print("READING RF–TEA/LCA WORKBOOK")
print("=" * 80)
print(INPUT_FILE)

comparison = pd.read_excel(
    INPUT_FILE,
    sheet_name=COMPARISON_SHEET,
)

predicted_lca = pd.read_excel(
    INPUT_FILE,
    sheet_name=PREDICTED_LCA_SHEET,
)

comparison.columns = (
    comparison.columns.astype(str).str.strip()
)

predicted_lca.columns = (
    predicted_lca.columns.astype(str).str.strip()
)

require_columns(
    comparison,
    ["Config", "Index"],
    COMPARISON_SHEET,
)


# ============================================================
# READ EXISTING SOLIDS-FREE YIELDS
# ============================================================

solids_free_yields = prepare_solids_free_yields(
    predicted_lca
)

base = comparison.merge(
    solids_free_yields,
    on=["Config", "Index"],
    how="inner",
    validate="one_to_one",
)

print(f"\nRows in comparison sheet: {len(comparison)}")
print(
    "Rows after solids-free yield merge: "
    f"{len(base)}"
)

# ============================================================
# REMOVE KNOWN NORMALIZATION ARTIFACTS
# ============================================================

before = len(base)

# Preserve excluded rows for the output workbook
artifact_rows = base.loc[
    base["Index"].isin(EXCLUDED_INDICES)
].copy()

# Remove artifacts from all subsequent analyses
base = base.loc[
    ~base["Index"].isin(EXCLUDED_INDICES)
].copy()

print(
    f"Removed {before - len(base)} known artifact rows "
    f"(Index {EXCLUDED_INDICES})"
)
print(f"Rows retained: {len(base)}")


# ============================================================
# CREATE NINE TRACI TABLES
# ============================================================

metric_sheets = {}

print("\n" + "=" * 80)
print("CREATING PAIRED METRIC TABLES")
print("=" * 80)

for settings in CATEGORY_SETTINGS:
    experimental_column = settings["experimental"]
    predicted_column = settings["predicted"]

    require_columns(
        comparison,
        [experimental_column, predicted_column],
        COMPARISON_SHEET,
    )

    table = create_metric_sheet(
        base=base,
        experimental_column=experimental_column,
        predicted_column=predicted_column,
        metric_label=settings["label"],
        metric_unit=settings["unit"],
    )

    metric_sheets[settings["sheet"]] = table

    print(
        f"{settings['sheet']}: "
        f"{len(table)} valid paired rows"
    )


# ============================================================
# CREATE IRR TABLE
# ============================================================

require_columns(
    comparison,
    [
        "Experimental IRR (%)",
        "Predicted IRR (%)",
    ],
    COMPARISON_SHEET,
)

irr_table = create_metric_sheet(
    base=base,
    experimental_column="Experimental IRR (%)",
    predicted_column="Predicted IRR (%)",
    metric_label="IRR",
    metric_unit="percentage points",
)

metric_sheets["IRR"] = irr_table

print(
    f"IRR: {len(irr_table)} valid paired rows"
)


# ============================================================
# CREATE OVERALL SUMMARY
# ============================================================

overall_summary = pd.DataFrame(
    [
        summarize_metric_table(sheet_name, table)
        for sheet_name, table in metric_sheets.items()
    ]
)


# ============================================================
# CREATE 20–50% BIOCRUDE FILTERED TABLES AND SUMMARY
# ============================================================

filtered_sheets = {}
filtered_summary_rows = []

for sheet_name, table in metric_sheets.items():
    filtered = table.loc[
        table["Biocrude (sfwt%)"].between(
            BIOCRUDE_MIN,
            BIOCRUDE_MAX,
            inclusive="both",
        )
    ].copy()

    filtered = filtered.reset_index(drop=True)

    filtered_sheets[sheet_name] = filtered

    filtered_summary_rows.append(
        summarize_filtered_table(
            sheet_name,
            filtered,
        )
    )

filtered_summary = pd.DataFrame(
    filtered_summary_rows
)


# ============================================================
# SAVE WORKBOOK 1: ALL PAIRED ROWS
# ============================================================

with pd.ExcelWriter(
    TERNARY_OUTPUT_FILE,
    engine="openpyxl",
) as writer:

    overall_summary.to_excel(
        writer,
        sheet_name="Summary",
        index=False,
    )

    # Save excluded rows for transparency
    artifact_rows.to_excel(
        writer,
        sheet_name="Excluded Artifacts",
        index=False,
    )

    for sheet_name, table in metric_sheets.items():
        table.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False,
        )

format_excel(TERNARY_OUTPUT_FILE)

# ============================================================
# SAVE WORKBOOK 2: 20–50% SOLIDS-FREE BIOCRUDE
# ============================================================

with pd.ExcelWriter(
    FILTERED_OUTPUT_FILE,
    engine="openpyxl",
) as writer:

    # Write the summary first so the workbook always has
    # at least one visible sheet.
    filtered_summary.to_excel(
        writer,
        sheet_name="Summary",
        index=False,
    )

    for sheet_name, table in filtered_sheets.items():
        table.to_excel(
            writer,
            sheet_name=sheet_name[:31],
            index=False,
        )

format_excel(FILTERED_OUTPUT_FILE)


# ============================================================
# FINAL REPORT
# ============================================================

print("\n" + "=" * 80)
print("COMPLETED")
print("=" * 80)

print(
    "\nAll paired ternary data saved to:\n"
    f"{TERNARY_OUTPUT_FILE}"
)

print(
    "\n20–50% solids-free biocrude data and summary saved to:\n"
    f"{FILTERED_OUTPUT_FILE}"
)

print("\n20–50% solids-free biocrude sample counts:")

for _, row in filtered_summary.iterrows():
    print(
        f"  {row['Metric']}: "
        f"{int(row['Valid samples'])}"
    )